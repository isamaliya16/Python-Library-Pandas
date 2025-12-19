
import pandas as pd
from openpyxl import load_workbook

# Load existing Excel file and append data
data = {'A': [7, 8, 9], 'B': [10, 11, 12]}
df = pd.DataFrame(data)
book = load_workbook('13 - Excel Handling/data/11_append.xlsx')
writer = pd.ExcelWriter('13 - Excel Handling/data/11_append.xlsx', engine='openpyxl')
writer.book = book
df.to_excel(writer, index=False, header=False, startrow=book['Sheet1'].max_row)
writer.save()
print('Data appended to 11_append.xlsx')
    
